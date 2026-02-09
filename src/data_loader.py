import openpyxl
from pathlib import Path
from datetime import datetime

class DataLoader:
    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

    def load_data(self):
        print(f"Loading data from: {self.excel_path.name}...")
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        # Initialize the context dictionary that will be passed to Jinja2
        context = {
            "generated_at": datetime.now().strftime("%Y-%m-%d"),
            "executive_summary": {},
            "district_performance": [],
            "financial_highlights": [],
            "staff_stats": [],
            "events": []
        }

        # 1. Parse Executive Summary (Key-Value lookup)
        if 'Executive_Summary' in wb.sheetnames:
            context['executive_summary'] = self._parse_executive_summary(wb['Executive_Summary'])
        
        # 2. Parse District Performance (Table)
        if 'District_Performance' in wb.sheetnames:
            context['district_performance'] = self._parse_district_performance(wb['District_Performance'])

        # 3. Parse Financial Highlights (Table)
        if 'Financial_Highlights' in wb.sheetnames:
            context['financial_highlights'] = self._parse_financial_highlights(wb['Financial_Highlights'])

        # 4. Parse Staff Stats (Table)
        if 'Staff_Statistics' in wb.sheetnames:
            context['staff_stats'] = self._parse_staff_stats(wb['Staff_Statistics'])
            
        # 5. Parse Monthly Events (Table)
        if 'Monthly_Events' in wb.sheetnames:
            context['events'] = self._parse_events(wb['Monthly_Events'])

        print("✓ Data successfully loaded.")
        return context

    def _parse_executive_summary(self, sheet):
        """
        Reads rows like: [Category, Metric_Name, Current_Value, Target_Value]
        Converts them into a dictionary for easy access: data['total_revenue']
        """
        data = {}
        # Iterate from row 2 to skip header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            metric_name = row[1]  # Column B
            value = row[2]        # Column C
            
            if not metric_name: continue

            # Map Excel names to Python keys
            if "Total Revenue" in str(metric_name):
                data['total_revenue'] = value
            elif "Member Collections" in str(metric_name):
                data['member_collections'] = value
            elif "Total New Enrollments" in str(metric_name):
                data['new_enrollments'] = value
            elif "Active Members" in str(metric_name):
                data['active_members'] = value
            elif "Pensioners" in str(metric_name):
                data['pensioners'] = value

        return data

    def _parse_district_performance(self, sheet):
        """
        Reads: [District, Recruitment_Count, Target_Count]
        Fix: Forces values to float/int to prevent 'str' division errors.
        """
        data = []
        # iter_rows returns tuples. row[0]=District, row[1]=Recruitment, row[2]=Target
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not row[0]: 
                continue

            try:
                # Force conversion to numbers (handle strings like "1,000" or "500")
                # We use str().replace(',', '') just in case Excel sent "1,200" as text
                recruitment = float(str(row[1]).replace(',', '')) if row[1] else 0
                target = float(str(row[2]).replace(',', '')) if row[2] else 0
                
                achievement = 0
                if target > 0:
                    achievement = round((recruitment / target) * 100, 1)

                data.append({
                    "district": row[0],
                    "recruitment": recruitment,
                    "target": target,
                    "achievement": achievement
                })
            except ValueError:
                print(f"⚠️ Warning: Could not parse numbers for district {row[0]}. Skipping.")
                continue
                
        return data

    def _parse_financial_highlights(self, sheet):
        """
        Reads: [Description, Amount_2026_Rs, Amount_2025_Rs]
        """
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append({
                    "description": row[0],
                    "amount_2026": row[1] or 0,
                    "amount_2025": row[2] or 0
                })
        return data

    def _parse_staff_stats(self, sheet):
        """
        Reads: [Designation, Approved_Cadre, Existing_Cadre, Vacancies]
        """
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append({
                    "designation": row[0],
                    "approved": row[1] or 0,
                    "existing": row[2] or 0,
                    "vacancies": row[3] or 0
                })
        return data

    def _parse_events(self, sheet):
        """
        Reads: [Date, Event_Title, Description_Sinhala, Image_Filename]
        """
        events = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                events.append({
                    "date": row[0], # datetime object from Excel
                    "title": row[1],
                    "description": row[2],
                    "image": row[3]
                })
        return events